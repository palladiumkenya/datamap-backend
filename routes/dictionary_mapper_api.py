from fastapi import  Depends, HTTPException

from sqlalchemy import create_engine, inspect, MetaData, Table,text
from sqlalchemy.orm import sessionmaker
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import sessionmaker, Session

import json
from fastapi import APIRouter
from typing import List
import pandas as pd
import requests
from openpyxl import load_workbook
from io import BytesIO
import logging


from models.models import AccessCredentials,IndicatorVariables
from database import database
from serializers.dictionary_mapper_serializer import indicator_selector_list_entity,indicator_list_entity
from serializers.access_credentials_serializer import access_credential_list_entity


log = logging.getLogger()
log.setLevel('DEBUG')
handler = logging.StreamHandler()
handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(name)s: %(message)s"))
log.addHandler(handler)




router = APIRouter()


# # Create an inspector object to inspect the database
engine = None
inspector = None

SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)


def createEngine():
    global engine, inspector, metadata
    try:
        credentials = AccessCredentials.objects().all()
        if credentials:
            credentials = access_credential_list_entity(credentials)
            connection_string = credentials
            # engine = create_engine(connection_string[0]["conn_string"])
            engine = create_engine(connection_string[0]["conn_string"])

            inspector = inspect(engine)
            metadata = MetaData()
            metadata.reflect(bind=engine)

    except SQLAlchemyError as e:
        # Log the error or handle it as needed
        raise HTTPException(status_code=500, detail="Database connection error")


def get_db():
    db_session = SessionLocal()
    try:
        yield db_session
    finally:
        db_session.close()

# @router.on_event("startup")
async def startup_event():
    createEngine()
    if engine:
        SessionLocal.configure(bind=engine)
    else:
        raise HTTPException(status_code=500, detail="Failed to initialize database engine")


@router.get('/base_schemas')
async def base_schemas():

    # Check if download was successful
    try:

        wb = load_workbook('configs/data_dictionary/dictionary.xlsx')
        sheets = wb.sheetnames
        print('sheets ', sheets)
        cass_session = database.cassandra_session_factory()

        schemas = []

        for schema in sheets:
            print('schema --> ', schema)

            df = pd.read_excel('configs/data_dictionary/dictionary.xlsx', sheet_name=schema)
            base_variables = []
            for i in range(0, df.shape[0]):
                query = "SELECT * FROM indicator_variables WHERE base_variable_mapped_to='%s' and base_repository='%s' ALLOW FILTERING;"%(df['Column/Variable Name'][i], schema)
                # results = database.execute_query(query)
                rows= cass_session.execute(query)
                results = []
                for row in rows:
                    results.append(row)
                matchedVariable = False if not results else True

                base_variables.append({'variable':df['Column/Variable Name'][i], 'matched':matchedVariable})

            baseSchemaObj = {}
            baseSchemaObj["schema"] = schema
            baseSchemaObj["base_variables"] = base_variables

            schemas.append(baseSchemaObj)
        return schemas
    except Exception as e:
        log.error('System ran into an error --> ', e)
        return e



@router.get('/base_schema_variables/{base_lookup}')
async def base_variables(base_lookup: str):
    try:

        cass_session = database.cassandra_session_factory()

        schemas = []

        df = pd.read_excel('configs/data_dictionary/dictionary.xlsx', sheet_name=base_lookup)
        base_variables = []
        for i in range(0, df.shape[0]):
            query = "SELECT * FROM indicator_variables WHERE base_variable_mapped_to='%s' and base_repository='%s' ALLOW FILTERING;"%(df['Column/Variable Name'][i], base_lookup)
            rows= cass_session.execute(query)
            results = []
            for row in rows:
                results.append(row)
            matchedVariable = False if not results else True

            base_variables.append({'variable':df['Column/Variable Name'][i], 'matched':matchedVariable})

        baseSchemaObj = {}
        baseSchemaObj["schema"] = base_lookup
        baseSchemaObj["base_variables"] = base_variables

        schemas.append(baseSchemaObj)
        return schemas
    except Exception as e:
        log.error('System ran into an error fetching base_schema_variables --->', e)
        return e


@router.get('/base_variables/{base_lookup}')
async def base_variables(base_lookup: str):

    try:
        df = pd.read_excel('configs/data_dictionary/dictionary.xlsx', sheet_name=base_lookup)

        base_variables = []
        for i in range(0, df.shape[0]):
            base_variables.append(df['Column/Variable Name'][i])
        return base_variables
    except Exception as e:
        log.error('System ran into an error fetching base_variables --->', e)
        return e





@router.get('/getDatabaseColumns')
async def getDatabaseColumns():
    try:
        dbTablesAndColumns={}

        table_names = metadata.tables.keys()

        for table_name in table_names:

            columns = inspector.get_columns(table_name)

            getcolumnnames = []
            for column in columns:
                getcolumnnames.append(column['name'])

            dbTablesAndColumns[table_name] = getcolumnnames
        # credential = credential
        # print("dbTablesAndColumns =======>",dbTablesAndColumns)
        return dbTablesAndColumns
    except SQLAlchemyError as e:
        log.error('Error reflecting database: --->', e)



@router.post('/add_mapped_variables')
async def add_mapped_variables(variables:List[object]):
    try:
        for variableSet in variables:
            IndicatorVariables.create(tablename=variableSet["tablename"],columnname=variableSet["columnname"],
                                                   datatype=variableSet["datatype"], base_repository=variableSet["base_repository"],
                                                   base_variable_mapped_to=variableSet["base_variable_mapped_to"])
        return {"status":200, "message":"Mapped Variables added"}
    except Exception as e:
        return {"status":500, "message":e}


@router.get('/tx_curr_variables')
async def available_connections():
    variables = IndicatorVariables.objects().all()
    variables = indicator_selector_list_entity(variables)

    # print(credentials)
    return {'variables': variables}



@router.get('/generate_config')
async def generate_config(baseSchema:str):
    access_cred = AccessCredentials.objects().all()
    access_cred = access_credential_list_entity(access_cred)

    config = IndicatorVariables.objects().all()
    config = indicator_selector_list_entity(config)
    with open('configs/schemas/'+baseSchema +'.conf', 'w') as f:
        f.write(str(config))

    return 'success'



@router.get('/import_config')
async def import_config(baseSchema:str):
    try:

        f = open('configs/schemas/'+baseSchema +'.conf', 'r')

        configImportStatements = f.read()
        configs = json.loads(configImportStatements.replace("'", '"'))
        for conf in configs:
            IndicatorVariables.create(tablename=conf["tablename"], columnname=conf["columnname"],
                                      datatype=conf["datatype"], base_repository=conf["base_repository"],
                                      base_variable_mapped_to=conf["base_variable_mapped_to"])

        f.close()
        log.info("========= Successfully imported config ==========")

        return 'success'
    except Exception as e:
        log.error("Error importing config ==> %s", str(e))

        return e


# @router.get('/generate-query/{baselookup}')
def generate_query(baselookup:str):
    try:

        cass_session = database.cassandra_session_factory()

        query = "SELECT * FROM indicator_variables WHERE base_repository='%s' ALLOW FILTERING;" % (baselookup)
        configs = cass_session.execute(query)

        mapped_columns = []
        mapped_joins = []

        for conf in configs:
            mapped_columns.append(conf["tablename"]+ "."+conf["columnname"] +" as '"+conf["base_variable_mapped_to"]+"' ")
            if all(conf["tablename"]+"." not in s for s in mapped_joins):

                mapped_joins.append(" LEFT JOIN "+conf["tablename"] + " ON " + "etl_patient_demographics.patient_id = " + conf["tablename"].strip() +".patient_id ")

        columns = ", ".join(mapped_columns)
        joins = ", ".join(mapped_joins)

        query = f"SELECT uuid() as id, {columns} from etl_patient_demographics {joins.replace(',','')} limit 100"

        log.info("========= Successfully generated query ==========")

        return query
    except Exception as e:
        log.error("Error importing config ==> %s", str(e))

        return e


@router.get('/load_data/{baselookup}')
async def load_data(baselookup:str, db_session: Session = Depends(get_db)):
    try:
        query = text(generate_query(baselookup))
        # query=text("select uuid() as id, hiv.*    from etl_hiv_enrollment hiv ")

        with db_session as session:
            result = session.execute(query)

            columns=result.keys()
            baseRepoLoaded = [dict(zip(columns,row)) for row in result]

            return baseRepoLoaded
    except Exception as e:
        log.error("Error loading data ==> %s", str(e))

        return e