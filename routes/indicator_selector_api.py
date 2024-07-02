from sqlalchemy import create_engine, inspect, MetaData, Table,text
from sqlalchemy.orm import sessionmaker

import json
from fastapi import APIRouter
from typing import List
import pandas as pd
import requests
from openpyxl import load_workbook
from io import BytesIO
import logging


from models.models import AccessCredentials,IndicatorVariables
from database import database
from serializers.indicator_selector_serializer import indicator_selector_list_entity,indicator_list_entity
from serializers.access_credentials_serializer import access_credential_list_entity


log = logging.getLogger()
log.setLevel('DEBUG')
handler = logging.StreamHandler()
handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(name)s: %(message)s"))
log.addHandler(handler)




router = APIRouter()



# Create a SQLite database engine
def get_connection_string():
    credentials = AccessCredentials.objects().all()
    if credentials:
        credentials = access_credential_list_entity(credentials)

        connection_string = credentials

        engine = create_engine(connection_string[0]["conn_string"])

        return engine


engine = get_connection_string()


# Create an inspector object to inspect the database
inspector = inspect(engine)
metadata = MetaData()
metadata.reflect(bind=engine)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)






@router.get('/base_schemas')
async def base_schemas():
    # URL of the Excel file
    excel_url = 'https://docs.google.com/spreadsheets/d/e/2PACX-1vRPNM8D6PJQPHjFur1f7QBE0x2B9HqOFzIkHQgwOcOQJlKR4EcHWCC5dP5Fm7MBlUN2G3QymZiu_xKy/pub?output=xlsx'
    # Download the file
    response = requests.get(excel_url)

    # Check if download was successful
    if response.status_code == 200:

        with open('configs/data_dictionary/dictionary.xlsx', 'wb') as f:
            f.write(response.content)
        print('File downloaded successfully')

        wb = load_workbook(BytesIO(response.content))
        sheets = wb.sheetnames

        cass_session = database.cassandra_session_factory()

        schemas = []

        for schema in sheets:
            df = pd.read_excel('configs/data_dictionary/dictionary.xlsx', sheet_name=schema)
            base_variables = []
            for i in range(0, df.shape[0]):
                query = "SELECT * FROM indicator_variables WHERE base_variable_mapped_to='%s' and indicator='%s' ALLOW FILTERING;"%(df['Column/Variable Name'][i], schema)
                # results = database.execute_query(query)
                rows= cass_session.execute(query)
                results = []
                for row in rows:
                    results.append(row)
                matchedVariable = False if not results else True

                base_variables.append({'variable':df['Column/Variable Name'][i], 'matched':matchedVariable})

            baseSchemaObj = {}
            baseSchemaObj["schema"] = schema
            baseSchemaObj["base_variables"] = base_variables

            schemas.append(baseSchemaObj)

    else:
        print('Failed to download file')

    return schemas



@router.get('/base_schema_variables/{base_lookup}')
async def base_variables(base_lookup: str):
    # URL of the Excel file
    excel_url = 'https://docs.google.com/spreadsheets/d/e/2PACX-1vRPNM8D6PJQPHjFur1f7QBE0x2B9HqOFzIkHQgwOcOQJlKR4EcHWCC5dP5Fm7MBlUN2G3QymZiu_xKy/pub?output=xlsx'
    # Download the file
    response = requests.get(excel_url)

    # Check if download was successful
    if response.status_code == 200:

        with open('configs/data_dictionary/dictionary.xlsx', 'wb') as f:
            f.write(response.content)
        print('File downloaded successfully')

        wb = load_workbook(BytesIO(response.content))
        sheets = wb.sheetnames

        cass_session = database.cassandra_session_factory()

        schemas = []

        df = pd.read_excel('configs/data_dictionary/dictionary.xlsx', sheet_name=base_lookup)
        base_variables = []
        for i in range(0, df.shape[0]):
            query = "SELECT * FROM indicator_variables WHERE base_variable_mapped_to='%s' and indicator='%s' ALLOW FILTERING;"%(df['Column/Variable Name'][i], base_lookup)
            # results = database.execute_query(query)
            rows= cass_session.execute(query)
            results = []
            for row in rows:
                results.append(row)
            matchedVariable = False if not results else True

            base_variables.append({'variable':df['Column/Variable Name'][i], 'matched':matchedVariable})

        baseSchemaObj = {}
        baseSchemaObj["schema"] = base_lookup
        baseSchemaObj["base_variables"] = base_variables

        schemas.append(baseSchemaObj)

    else:
        print('Failed to download file')

    return schemas


@router.get('/base_variables/{base_lookup}')
async def base_variables(base_lookup: str):
    # URL of the Excel file
    excel_url = 'https://docs.google.com/spreadsheets/d/e/2PACX-1vRPNM8D6PJQPHjFur1f7QBE0x2B9HqOFzIkHQgwOcOQJlKR4EcHWCC5dP5Fm7MBlUN2G3QymZiu_xKy/pub?output=xlsx'

    # Download the file
    response = requests.get(excel_url)

    # Check if download was successful
    if response.status_code == 200:
        with open('configs/data_dictionary/dictionary.xlsx', 'wb') as f:
            f.write(response.content)
        print('File downloaded successfully')
        df = pd.read_excel('configs/data_dictionary/dictionary.xlsx', sheet_name=base_lookup)

        # Display the DataFrame
        print(df.head())

        base_variables = []
        for i in range(0, df.shape[0]):
            base_variables.append(df['Column/Variable Name'][i])
        print('base_variables===>',base_variables)
    else:
        print('Failed to download file')
    print('base_lookup',base_lookup)

    return base_variables



@router.get('/getDatabaseColumns')
async def getDatabaseColumns():

    dbTablesAndColumns={}

    table_names = metadata.tables.keys()
    for table_name in table_names:

        columns = inspector.get_columns(table_name)

        getcolumnnames = []
        for column in columns:
            getcolumnnames.append(column['name'])

        dbTablesAndColumns[table_name] = getcolumnnames
    # credential = credential
    return dbTablesAndColumns


@router.post('/add_mapped_variables')
async def add_mapped_variables(variables:List[object]):
    try:
        for variableSet in variables:
            IndicatorVariables.create(tablename=variableSet["tablename"],columnname=variableSet["columnname"],
                                                   datatype=variableSet["datatype"], indicator=variableSet["indicator"],
                                                   base_variable_mapped_to=variableSet["base_variable_mapped_to"])
        return {"status":200, "message":"Mapped Variables added"}
    except Exception as e:
        return {"status":500, "message":e}


@router.get('/tx_curr_variables')
async def available_connections():
    variables = IndicatorVariables.objects().all()
    variables = indicator_selector_list_entity(variables)

    # print(credentials)
    return {'variables': variables}


@router.get('/tx_curr_generate_indicator')
async def generate_indicator(indicator:str):
    print("passed indicator", indicator)
    query="select 'TX_CURR'                                                                   AS 'indicator', " \
                "       count(distinct t.patient_id)                                                as 'indicator_value', " \
                "       date_format(last_day(date_sub(current_date(), interval 1 MONTH)), '%Y%b%d') as 'indicator_date' " \
                "from( " \
                "    select fup.visit_date,fup.patient_id, max(e.visit_date) as enroll_date, " \
                "           greatest(max(e.visit_date), ifnull(max(date(e.transfer_in_date)),'0000-00-00')) as latest_enrolment_date, " \
                "           greatest(max(fup.visit_date), ifnull(max(d.visit_date),'0000-00-00')) as latest_vis_date, " \
                "           greatest(mid(max(concat(fup.visit_date,fup.next_appointment_date)),11), ifnull(max(d.visit_date),'0000-00-00')) as latest_tca, " \
                "           d.patient_id as disc_patient, " \
                "           d.effective_disc_date as effective_disc_date, " \
                "           max(d.visit_date) as date_discontinued, " \
                "           de.patient_id as started_on_drugs " \
                "    from kenyaemr_etl.etl_patient_hiv_followup fup " \
                "           join kenyaemr_etl.etl_patient_demographics p on p.patient_id=fup.patient_id " \
                "           join kenyaemr_etl.etl_hiv_enrollment e on fup.patient_id=e.patient_id " \
                "           left outer join kenyaemr_etl.etl_drug_event de on e.patient_id = de.patient_id and de.program='HIV' and date(date_started) <= date(date(last_day(date_sub(current_date(),interval 1 MONTH)))) " \
                "           left outer JOIN " \
                "             (select patient_id, coalesce(date(effective_discontinuation_date),visit_date) visit_date,max(date(effective_discontinuation_date)) as effective_disc_date from kenyaemr_etl.etl_patient_program_discontinuation " \
                "              where date(visit_date) <= date(date(last_day(date_sub(current_date(),interval 1 MONTH)))) " \
                "                and program_name='HIV' and patient_id " \
                "              group by patient_id " \
                "             ) d on d.patient_id = fup.patient_id " \
                "    where fup.visit_date <= date(date(last_day(date_sub(current_date(),interval 1 MONTH)))) " \
                "    group by patient_id " \
                "    having (started_on_drugs is not null and started_on_drugs <> '') " \
                "       and " \
                "           ( " \
                "               ((timestampdiff(DAY,date(latest_tca),date(date(last_day(date_sub(current_date(),interval 1 MONTH))))) <= 30 or timestampdiff(DAY,date(latest_tca),date(curdate())) <= 30) and " \
                "                ((date(d.effective_disc_date) > date(date(last_day(date_sub(current_date(),interval 1 MONTH)))) or date(enroll_date) > date(d.effective_disc_date)) or d.effective_disc_date is null)) " \
                "                 and " \
                "               (date(latest_vis_date) >= date(date_discontinued) or date(latest_tca) >= date(date_discontinued) or disc_patient is null) " \
                "               ) " \
                "    )t; "

    with SessionLocal() as session:
        result = session.execute(query)
        indicatorRes = [dict(row) for row in result]

        indicators = indicator_list_entity(indicatorRes)
        session.shutdown()

        return {"indicators": indicators}



@router.get('/generate_config')
async def generate_config(baseSchema:str):
    access_cred = AccessCredentials.objects().all()
    access_cred = access_credential_list_entity(access_cred)

    config = IndicatorVariables.objects().all()
    config = indicator_selector_list_entity(config)
    with open('configs/schemas/'+baseSchema +'.conf', 'w') as f:
        f.write(str(config))

    return 'success'



@router.get('/import_config')
async def import_config(baseSchema:str):
    try:
        access_cred = AccessCredentials.objects().all()
        access_cred = access_credential_list_entity(access_cred)

        f = open('configs/schemas/'+baseSchema +'.conf', 'r')

        configImportStatements = f.read()
        configs = json.loads(configImportStatements.replace("'", '"'))
        for conf in configs:
            IndicatorVariables.create(tablename=conf["tablename"], columnname=conf["columnname"],
                                      datatype=conf["datatype"], indicator=conf["indicator"],
                                      base_variable_mapped_to=conf["base_variable_mapped_to"])


        f.close()
        log.info("========= Successfully imported config ==========")

        return 'success'
    except Exception as e:
        log.error("Error importing config ==> %s", str(e.message))

        return e


@router.get('/generate-query')
async def generate_query():
    try:
        access_cred = AccessCredentials.objects().all()
        access_cred = access_credential_list_entity(access_cred)

        configs = IndicatorVariables.objects().all()
        configs = indicator_selector_list_entity(configs)

        mapped_columns = []
        for conf in configs:
            mapped_columns.append(conf["tablename"]+ "."+conf["columnname"])
        print("mapped_columns -> ",mapped_columns)
        columns = ", ".join(mapped_columns)

        query = f"SELECT {columns} from ..."
        print("query -> ",query)

        log.info("========= Successfully generated query ==========")

        return 'success'
    except Exception as e:
        log.error("Error importing config ==> %s", str(e))

        return e



